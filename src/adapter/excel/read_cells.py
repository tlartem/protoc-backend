from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def column_number_to_letter(column_number: int) -> str:
    """Конвертирует номер столбца в буквенное обозначение Excel (1 -> A, 2 -> B, 27 -> AA, etc.)"""
    result = ""
    while column_number > 0:
        column_number -= 1
        result = chr(65 + column_number % 26) + result
        column_number //= 26
    return result


def read_cells(file_content: bytes) -> dict[str, Any]:
    # Создаем BytesIO объект из байтов файла
    file_stream = BytesIO(file_content)
    try:
        workbook = load_workbook(file_stream, data_only=True)
        sheet = workbook.active
    except InvalidFileException:
        raise ValueError("Файл не является валидным Excel файлом (.xlsx, .xlsm, .xltx, .xltm)")

    cells = {}
    merged = []
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells[cell.coordinate] = cell.value
    
    for merged_range in sheet.merged_cells.ranges:
        merged.append({
            "from": f"{column_number_to_letter(merged_range.min_col)}{merged_range.min_row}",
            "to": f"{column_number_to_letter(merged_range.max_col)}{merged_range.max_row}"
        })

    return {
        "cells": cells,
        "merged": merged
    }